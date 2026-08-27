import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
OUTPUT = os.path.join(os.path.dirname(__file__), "latam_transit_boards.xlsx")

CATEGORIES = [
    "Transit Ops/Management",
    "Other Management/Policy",
    "Labor Representative",
    "Community Advocate",
    "Elected Official",
]

CATEGORY_COLORS = {
    "Transit Ops/Management": "1F4E79",
    "Other Management/Policy": "2E75B6",
    "Labor Representative": "BF8F00",
    "Community Advocate": "548235",
    "Elected Official": "C00000",
}

CATEGORY_FONTS = {
    "Transit Ops/Management": "FFFFFF",
    "Other Management/Policy": "FFFFFF",
    "Labor Representative": "FFFFFF",
    "Community Advocate": "FFFFFF",
    "Elected Official": "FFFFFF",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def load_agencies():
    agencies = []
    for fname in sorted(os.listdir(DATA_DIR)):
        if fname.endswith(".json"):
            with open(os.path.join(DATA_DIR, fname)) as f:
                agencies.append(json.load(f))
    return agencies

def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, wrap=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def create_board_members_sheet(wb, agencies):
    ws = wb.active
    ws.title = "Board Members"

    headers = [
        "Agency", "City", "Country", "Member Name", "Position",
        "Appointment Method", "Professional Background", "Education",
        "Day Classification", "Judgment Call?", "Why This Call",
        "Classification Rationale", "Source URL(s)",
        "Confidence Level", "Date Verified"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for agency in agencies:
        for m in agency["members"]:
            sources = "; ".join(m.get("sources", [])) if isinstance(m.get("sources"), list) else m.get("sources", "")
            values = [
                agency["agency"], agency["city"], agency["country"],
                m["name"], m["position"], m.get("appointment_method", ""),
                m.get("background", ""), m.get("education", ""),
                m["classification"],
                "Yes" if m.get("judgment_call") else "",
                m.get("classification_note", ""),
                m.get("rationale", ""),
                sources, m.get("confidence", ""), agency.get("date_verified", "")
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                style_data_cell(cell, wrap=(col in (5, 6, 7, 8, 11, 12, 13)))

            cls_cell = ws.cell(row=row, column=9)
            cls = m["classification"]
            if cls in CATEGORY_COLORS:
                cls_cell.fill = PatternFill("solid", fgColor=CATEGORY_COLORS[cls])
                cls_cell.font = Font(color=CATEGORY_FONTS[cls], bold=True)

            row += 1

    col_widths = [35, 14, 12, 28, 32, 30, 50, 35, 24, 50, 60, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "A2"

def create_agency_summary_sheet(wb, agencies):
    ws = wb.create_sheet("Agency Summary")

    headers = [
        "Agency", "City", "Country", "Board Size",
        "% Transit Ops/Mgmt", "% Other Mgmt/Policy", "% Labor Rep",
        "% Community Advocate", "% Elected Official",
        "Governance Model", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for agency in agencies:
        members = agency["members"]
        total = len(members)
        counts = {c: 0 for c in CATEGORIES}
        for m in members:
            cls = m["classification"]
            if cls in counts:
                counts[cls] += 1

        pcts = {c: counts[c] / total if total > 0 else 0 for c in CATEGORIES}

        values = [
            agency["agency"], agency["city"], agency["country"], total,
            pcts["Transit Ops/Management"],
            pcts["Other Management/Policy"],
            pcts["Labor Representative"],
            pcts["Community Advocate"],
            pcts["Elected Official"],
            agency.get("governance_model", ""),
            agency.get("notes", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_data_cell(cell, wrap=(col >= 10))
            if 5 <= col <= 9:
                cell.number_format = "0%"
                cat_idx = col - 5
                cat = CATEGORIES[cat_idx]
                if val > 0:
                    cell.fill = PatternFill("solid", fgColor=CATEGORY_COLORS[cat])
                    cell.font = Font(color="FFFFFF", bold=True)

        row += 1

    col_widths = [40, 14, 12, 12, 18, 18, 14, 18, 18, 55, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

def main():
    agencies = load_agencies()
    wb = Workbook()
    create_board_members_sheet(wb, agencies)
    create_agency_summary_sheet(wb, agencies)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")
    print(f"  {sum(len(a['members']) for a in agencies)} board members across {len(agencies)} agencies")

if __name__ == "__main__":
    main()
